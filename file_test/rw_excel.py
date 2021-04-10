# -*- coding:utf-8 -*-
# author:
# datetime:2019/8/27 16:25
import os
import xlrd, xlwt
import openpyxl


# 读excel
# wb = xlrd.open_workbook("excel.xlsx")
# sheet = wb.sheet_by_index(0)
# nrows = sheet.nrows
# for i in range(nrows):
#     print(sheet.row_values(i))
#
# # 写excel
# wt = xlwt.Workbook("excel.xlsx")
# sheet = wt.add_sheet("name")
# sheet.write("行", "列", "data")


# 写excel
# value3 = [["姓名", "性别", "年龄", "城市", "职业"],
#           ["111", "女", "66", "石家庄", "运维工程师"],
#           ["222", "男", "55", "南京", "饭店老板"],
#           ["333", "女", "27", "苏州", "保安"],]
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    # workbook.create_sheet(sheet_name)
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


# def write_excel_xls(path, sheet_name, value):
#     index = len(value)  # 获取需要写入数据的行数
#     workbook = xlwt.Workbook()  # 新建一个工作簿
#     sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
#     for i in range(0, index):
#         for j in range(0, len(value[i])):
#             sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）
#     workbook.save(path)  # 保存工作簿
#     print("xls格式表格写入数据成功！")




# write_excel_xlsx(r"H:\人脸优选\test.xlsx", "11", value3)



# 读excel
# wb = xlrd.open_workbook(r"F:\WXWork\1688850573609985\Cache\File\2020-11\反动组织P0-10月(2).xlsx")
# path = r'F:\WXWork\1688850573609985\Cache\File\2020-11\反动组织'
# sheet = wb.sheet_by_name('Sheet1')
# nrows = sheet.nrows
# for i in range(nrows):
#     print(sheet.row_values(i)[2])
#     if not os.path.exists(os.path.join(path, sheet.row_values(i)[2])):
#         os.makedirs(os.path.join(path, sheet.row_values(i)[2]))